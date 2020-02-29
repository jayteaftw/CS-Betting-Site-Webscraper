from selenium import webdriver
from time import sleep
from tqdm import tqdm
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
from openpyxl import workbook
from openpyxl import load_workbook
from threading import Thread
import sys

class WebCsBetBot():
    data = []
    def __init__(self):
        option = webdriver.ChromeOptions()
        chrome_prefs = {}
        option.experimental_options["prefs"] = chrome_prefs
        chrome_prefs["profile.default_content_settings"] = {"images": 2}
        chrome_prefs["profile.managed_default_content_settings"] = {"images": 2}
        self.driver = webdriver.Chrome(chrome_options=option)

        #self.driver = webdriver.Chrome()
    
    def __del__(self):
        self.driver.close()

    def GetWebPage(self):
        self.driver.get('https://csgoempire.com/history?seed=1852')
        sleep(2)

    def StoreData(self):
        i = 1
        final = 4500
        date = self.driver.find_element_by_xpath('//*[@id="page-scroll"]/div/section/div/h3').text[10:]
        os.chdir('/home/jaytea/Code Workspace/Python/CSBettingSite/data')
        if(date == ''):
            return
        f = open(date,"a+")
        print("Scraping "+ date)
        pbar = tqdm(i ,total= final, unit=" Bets")
        while(i <= final):
            try: 
                x = self.driver.find_element_by_xpath('//*[@id="page-scroll"]/div/section/div/div/div['+str(i)+']/span').text[10:]
                self.data.append(x)
                f.write(x+"\n")
            except :
                #i = final
                break
            finally:
                pbar.update()
            i+=1
        f.close
        pbar.close()
        print("Elements in " + str(date) + ": " + str(i- 1))
    
    
    def pagecycle(self, start = 1, stop = 1, incr = 1 ):
        i = start
        page = 0
        while( i <= stop):
            print("getting ", str(i))
            self.driver.get('https://csgoempire.com/history?seed='+str(i))
            try:
                element_present = EC.presence_of_element_located((By.XPATH, '//*[@id="page-scroll"]/div/section/div/div/div[0]/span'))
                WebDriverWait(self.driver, 3).until(element_present)
            except: TimeoutError
  
            self.StoreData()
            i += incr
            page += 1
            #print("Completed ", str(page)," pages out of ", str(stop - start))
        print("Total elements added " + str(len(self.data)))
          

    def bet1(self, bet = 0, bank = 0):
        train = 0
        maxtrain = 0
        OrBank = bank
        betTmp = bet
        negsum = 0 
        flag = "no"
        f = open('bet_results_test.txt',"r+")       
        for i in f.readlines():
            if(int(i[:-1]) > 6):
                bank += betTmp
                betTmp = bet
                train = 0
                negsum = 0
            else:
                bank -= betTmp
                negsum -= betTmp
                betTmp = betTmp*2
                train += 1
            if(train > maxtrain):
                print(negsum)
                maxtrain = train
            if(bank < 0):
                print("Bank went negative after " + str(train))
                return
            
            if((OrBank + negsum) < 0):
                flag = " yes"

        print("Final Sum: " + str(bank) + ", Profit: " + str(bank - OrBank) + ", Largest Train = " + str(maxtrain) + ", Possible Neg bank: " + flag)

    def bet1withlimit(self, bet = 0, bank = 0, limit = 0):
        train = 0
        maxtrain = 0
        OrBank = bank
        betTmp = bet
        negsum = 0 
        flag = "no"
        f = open('bet_results_test.txt',"r+")       
        for i in f.readlines():
            if(int(i[:-1]) > 6):
                bank += betTmp
                betTmp = bet
                train = 0
                negsum = 0
            else:
                bank -= betTmp
                negsum -= betTmp
                betTmp = betTmp*2
                train += 1
            if(train > maxtrain):
                maxtrain = train
            if(bank < 0):
                print("Bank went negative after " + str(train))
                return
            if(train == limit):
                betTmp = bet
            if((OrBank + negsum) < 0):
                flag = " yes"

        print("Final Sum: " + str(bank) + ", Profit: " + str(bank - OrBank) + ", Largest Train = " + str(maxtrain) + ", Possible Neg bank: " + flag)
        workbook = load_workbook(filename="sample.xlsx")
        sheet = workbook.active
        sheet.cell(row = 1, column = 1).value = "hello"


if __name__ == '__main__':
    bot1 = WebCsBetBot()
    print("Arg 1 ",sys.argv[1]," Arg 2 ",sys.argv[2], " Arg 3 ", sys.argv[3]  )
    bot1.pagecycle(int(sys.argv[1]), int(sys.argv[2]), int(sys.argv[3]))



'''for j in range(10):
    for i in range(5):
        print(str((j)*10 + 25) + ' bank, ' + str((i+1)*.01) + ' bet')
        bot.bet1withlimit(0.01*(i+1),25 +(j*10), 9)
        print(" ")
    print("--------------------------------------")'''
    
